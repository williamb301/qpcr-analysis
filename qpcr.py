import numpy as np
import pandas as pd
from matplotlib import pyplot as plt
from sklearn.linear_model import LinearRegression
import openpyxl
import sys


"""
standard_curve_plot : function

parameters : avg_CT list of avg CT per triplet 

creates a pyplot of log_copies vs avg CT values that are linearized
through a scikit LinearRegression model

creates an excel sheet and appends the pyplot along with slope/intercept values

returns slope, intercept, r2 of CT vs log copies plot
"""
def standard_curve_plot(avg_CT):
    CT = [i for i in avg_CT if avg_CT.index(i) % 4 == 0] #extrapolating std's averages
    del CT[-1] #remove 8th value, since std goes from 1-7
    CT.reverse()
    log_copies = [4,5,6,7,8,9,10]

    
    x = np.array(log_copies).reshape((-1,1))
    y = np.array(CT)
    model = LinearRegression().fit(x, y) #creates line of best fit for CT and log values
    CT_lin = model.predict(x) #obtains linear CT y values
    slope = float(model.coef_)
    intercept = model.intercept_
    r2 = model.score(x,y)
    

    plt.figure(figsize=(4,4)) #creating plot with pyplot
    plt.plot(log_copies,CT_lin,ls = '--')
    plt.scatter(log_copies, CT, marker = 'o')
    plt.title('Viral Titer Standard Curve')
    plt.xlabel('Log copies/ml') ; plt.ylabel('CT')
    plt.annotate('y = {}x + {}'.format(round(slope,4),round(intercept,4)), xy=(7, 1.5*CT[4]),size = 8)
    plt.annotate('r^2 = {}'.format(round(r2,4)), xy=(7, 1.5*CT[4]-1),size = 8)
    plt.savefig('std_c.png',dpi=150)

    wb = openpyxl.Workbook()
    sh = wb.create_sheet('graph')
    img = openpyxl.drawing.image.Image('std_c.png')
    sh.add_image(img, 'A8')
    plot_col = ['slope = ', str(slope), 'intercept = ', str(intercept), 'r^2 = ', str(r2)]
    sh.append(plot_col)
    avgct = ['Avg CT = ']
    for i in CT:
        avgct.append(i)
    sh.append(avgct)
    del wb['Sheet']
    wb.save(sys.argv[2])
    wb.close()
    return slope, intercept, r2 #returns slope and intercept values
    

"""
create_frame : function

parameters : table dataframe (excel sheet 1) raw data dataframe (excel sheet 2)

while creating columns for frame, calls standard_curve_plot function to 
retrieve slope and intercept information necessary for further calculations

returns frame : a dataframe with all calculations and results
"""
def create_frame(table,data):
    well96 = [] #list of all tests done in loading table 
    well_pos = [] #creating list for 'Well Position' column
    del_list = [] # blank's index values to be removed from CT list later
    for i in range(8): #i from 0-7 : A-H
        row = [] 
        for j in range(0,4): # j from 1-4 : '1-3' - '10-12'
            index = table.iloc[i,j+1] #name of test at [Letter(A-H),triplet]
            row.append(index)
            for q in range(1,4):
                well_pos.append('{}{}'.format(chr(65+i),q+3*j)) #adding 8 4 triplets of well positions
            if pd.isna(index): 
                for q in range(1,4):
                    del_list.append([chr(65+i),q+3*j]) #add blank's well letters/nums to del list
        well96.append(row)
    
    CT_vals = data.iloc[43:,8].tolist() #splicing CT data from raw data excel

    CT =[]
    for i in CT_vals:
        if type(i) == float:
            CT.append(i)
        else:
            CT.append(0) #keep CT at size 96 but any 'undetermined' vals are set to 0

    #change CT from list to list of lists of triplets to make average and removal easier
    CT_t = []
    for i in range(int(len(CT)/3)):
        triplet = []
        triplet.append(CT[3*i])
        triplet.append(CT[3*i+1])
        triplet.append(CT[3*i+2])
        CT_t.append(triplet)

    avg_CT = [] #creating list of average CT for each triplet 
    r_vals = [] #list of removed values to then highlight
    for i in CT_t:
        avg = sum(i)/3
        for j in i:
            if j >= (avg + 0.5) or j <= (avg - 0.5): 
                #then the value is an outlier and we remove it and find new average
                r_vals.append(j) #obtaining index of removed value
                i.remove(j)
        avg_CT.append(sum(i)/len(i))
        avg_CT.append('')
        avg_CT.append('')
    
    slope, intercept, r2 = standard_curve_plot(avg_CT) #create viral titer standard curve plot and get slope/intercept values
    
    log_copies = [(x-intercept)/slope for x in CT] #creates list of log copies by applying formula on each CT value
    
    copies_ml = [np.power(10,x) for x in log_copies] #creates list of copies/ml by applying formula on each log copy

    qpcr_titer = [x*4520*2 for x in copies_ml]

    qpcr_t = []  #creating list of triplets for qpcr_titer
    for i in range(int(len(qpcr_titer)/3)):
        triplet = []
        triplet.append(qpcr_titer[3*i])
        triplet.append(qpcr_titer[3*i+1])
        triplet.append(qpcr_titer[3*i+2])
        qpcr_t.append(triplet)
     
    
    avg_qpcr = [] #creating list of average qpcr for each triplet 
    sd_qpcr = [] #creating list of standard deviations for each triplet
    avg_qpcr2 = [] #for frame2
    sd_qpcr2 = [] #for frame2
    for i in qpcr_t:
        avg = sum(i)/3
        sd_qpcr.append(np.std(i))
        for j in i:
            if j > (avg + 0.5) or j < (avg - 0.5): 
                #then the value is an outlier and we remove it and find new average
                i.remove(j)
        avg_qpcr.append(sum(i)/len(i))
        avg_qpcr.append('')
        avg_qpcr.append('')
        sd_qpcr.append('')
        sd_qpcr.append('')

    #creating second pandas dataframe for results page 2
    useful_tests = [x for x in well_pos]
    avg_qpcr2 = [x for x in avg_qpcr]
    sd_qpcr2 = [x for x in sd_qpcr]

    #remove values by double for loop that goes through deleted wells list and 96->0
    #index based on letter and number (find formula), use for all columns
    del_list.reverse() #to prevent indexes from shifting when deleting early ones in list
    for del_wel in del_list:
        dindex = (ord(del_wel[0])-65)*12 + del_wel[1] -1
        for well in well_pos:
            if dindex == well_pos.index(well):
                del well_pos[dindex]
                del CT[dindex]
                del avg_CT[dindex]
                del log_copies[dindex]
                del copies_ml[dindex]
                del qpcr_titer[dindex]
                del avg_qpcr[dindex]
                del sd_qpcr[dindex]
                del useful_tests[dindex]
                del avg_qpcr2[dindex]
                del sd_qpcr2[dindex]

    del_vals = []
    for r in r_vals:
        if r in CT:
            del_vals.append(CT.index(r))

    #removing stds from frame2 columns
    #note: modifying a list while iterating over it can cause weird behavior
    for i in range(len(useful_tests),0,-1):
        well = useful_tests[i-1]
        n = well[1] #leaving out any std test which would be in columns 1-3
        if well[0] == 'H':
            continue #to keep DC
        if len(well) == 3: 
            continue #for 10,11,12
        if n == '1' or (n == '2' or n == '3'):
            dindex = useful_tests.index(well)
            useful_tests.remove(well)
            del avg_qpcr2[dindex]
            del sd_qpcr2[dindex]

    #converting to test_names from well positions for cleaner results page 2
    c1_3 = table['1-3'].to_list()
    c4_6 = table['4-6'].to_list()
    c7_9 = table['7-9'].to_list()
    c10_12 = table['10-12'].to_list()
    test_names = c1_3 + c4_6 + c7_9 + c10_12
    names = []
    for i in range(len(useful_tests)):
        well = useful_tests[i]
        dindex = ord(well[0])-65 
        col = int(well[1])
        if len(well) == 3:
            dindex += 24
        else: 
            if col >= 4 and col <= 6:
                dindex += 8
            elif col >= 7 and col <= 9:
                dindex += 16
        names.append(test_names[dindex])
    for i in range(len(names),0,-1):
        if i % 3 != 0:
            del names[i]
            del avg_qpcr2[i]
            del sd_qpcr2[i]


    frame = pd.DataFrame(well_pos,columns=['Well Position']) #creating dataframe with first column as well position
    frame['CT'] = CT #adding CT column to dataframe
    frame['Average CT'] = avg_CT #adding average CT column
    frame['log copies/ml'] = log_copies
    frame['copies/ml'] = copies_ml
    frame['qpcr titer'] = qpcr_titer
    frame['Average qpcr titer'] = avg_qpcr #adding average CT column
    frame['qpcr SD'] = sd_qpcr #adding standard deviation column

    #creating 2nd dataframe for 2nd results sheet
    frame2 = pd.DataFrame(names, columns = ['test names'])
    frame2['Average qpcr titer'] = avg_qpcr2
    frame2['qpcr SD'] = sd_qpcr2

    return frame, del_vals, frame2

#main handles command line inputs and adds frame dataframe to excel with pyplot
if __name__ == "__main__":
    if (len(sys.argv) != 3):
        print("No file input, argument format: './qpcr <input_file_name.(xlsx, xls, xlsm)> <output_file_name.(xlsx, xls, xlsm)>'")
    filename = sys.argv[1]
    data = pd.read_excel(filename, sheet_name = 0)
    loading_table = pd.read_excel(filename, sheet_name = 1)
    frame, del_vals, frame2 = create_frame(loading_table,data)
    with pd.ExcelWriter(sys.argv[2], mode='a', engine='openpyxl') as writer:  
        frame.to_excel(writer, sheet_name='results', index = False) #add results to existing excel sheet
    
    with pd.ExcelWriter(sys.argv[2], mode='a', engine='openpyxl') as writer:  
        frame2.to_excel(writer, sheet_name='results page 2', index = False) #add results to existing excel sheet

    wb = openpyxl.load_workbook(sys.argv[2]) #highlighting outlier cells
    ws = wb['results']
    yellow = "00FFFF00"
    for i in del_vals:
        highlight = ws.cell(row = i+2, column = 2) 
        highlight.fill = openpyxl.styles.PatternFill(start_color=yellow, end_color=yellow,fill_type = "solid")
    wb.save(sys.argv[2])

    